#!/usr/bin/env python3
"""
Take customers from Excel and fill shipping addresses from both address books.
Output: XLSX with customers + matched shipping addresses.
"""

import re
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def normalize_phone(phone: str) -> str:
    """Normalize phone for matching (last 9 digits)."""
    if not phone:
        return ""
    digits = re.sub(r'[^\d]', '', str(phone))
    return digits[-9:] if len(digits) >= 9 else digits

def normalize_name(name: str) -> str:
    """Normalize name for matching."""
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(name).lower())

def parse_csv_addressbook(filepath: str) -> list:
    """Parse the CSV address book."""
    addresses = []
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        lines = f.readlines()
    
    if not lines:
        return addresses
    
    headers = [h.strip() for h in lines[0].split(';')]
    
    for line in lines[1:]:
        if not line.strip():
            continue
        values = line.split(';')
        row = dict(zip(headers, [v.strip() for v in values]))
        
        country_map = {'CZ': 'Czech Republic', 'DE': 'Germany', 'AT': 'Austria', 
                      'SK': 'Slovakia', 'PL': 'Poland', 'DK': 'Denmark', 'BE': 'Belgium'}
        country = row.get('ZEME', 'CZ')
        
        addresses.append({
            'company': row.get('NAZEV', ''),
            'street': row.get('ULICE_CP', ''),
            'city': row.get('MESTO', ''),
            'zipCode': row.get('PSC', ''),
            'country': country_map.get(country, country),
            'email': row.get('EMAIL', ''),
            'contact': row.get('KONTAKTNI_OS', ''),
            'phone': row.get('TELEFON', ''),
            'phone_normalized': normalize_phone(row.get('TELEFON', '')),
            'company_normalized': normalize_name(row.get('NAZEV', '')),
            'contact_normalized': normalize_name(row.get('KONTAKTNI_OS', '')),
        })
    
    return addresses

def parse_german_addressbook(filepath: str) -> list:
    """Parse the German TXT address book."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    entries = re.split(r'\nBearbeiten\s*\n+Löschen\s*\n?', content)
    addresses = []
    
    for entry in entries:
        lines = [l.strip() for l in entry.strip().split('\n') if l.strip()]
        if len(lines) < 3:
            continue
        
        address = {
            'company': '',
            'contact': '',
            'street': '',
            'city': '',
            'zipCode': '',
            'country': 'Germany',
            'email': '',
            'phone': '',
        }
        
        for line in lines:
            if '@' in line:
                address['email'] = line
                break
        
        postal_idx = None
        for i, line in enumerate(lines):
            match = re.match(r'^(\d{4,5})\s+(.+?),\s*(\w{2})$', line)
            if match:
                address['zipCode'] = match.group(1)
                address['city'] = match.group(2)
                country_map = {'DE': 'Germany', 'AT': 'Austria', 'NL': 'Netherlands',
                              'DK': 'Denmark', 'CH': 'Switzerland', 'BE': 'Belgium', 'FR': 'France'}
                address['country'] = country_map.get(match.group(3).upper(), match.group(3))
                postal_idx = i
                break
        
        if postal_idx is None or postal_idx < 1:
            continue
        
        address['street'] = lines[postal_idx - 1]
        
        remaining = lines[:postal_idx - 1]
        if remaining:
            company_keywords = ['nails', 'beauty', 'studio', 'salon', 'center', 'shop', 'lounge', 'spa']
            first = remaining[0]
            if any(kw in first.lower() for kw in company_keywords):
                address['company'] = first
                if len(remaining) > 1 and not remaining[1].startswith('('):
                    address['contact'] = remaining[1]
            else:
                address['contact'] = first
                if len(remaining) > 1:
                    address['company'] = remaining[1]
        
        if address['street'] and address['zipCode']:
            address['phone_normalized'] = normalize_phone(address.get('phone', ''))
            address['company_normalized'] = normalize_name(address.get('company', ''))
            address['contact_normalized'] = normalize_name(address.get('contact', ''))
            addresses.append(address)
    
    return addresses

def find_matching_address(customer: dict, all_addresses: list) -> dict | None:
    """Find matching address for a customer."""
    cust_phone = normalize_phone(customer.get('phone', ''))
    cust_name = normalize_name(customer.get('name', ''))
    
    # Try phone match first (most reliable)
    if cust_phone:
        for addr in all_addresses:
            if addr['phone_normalized'] == cust_phone:
                return addr
    
    # Try company/contact name match
    if cust_name:
        for addr in all_addresses:
            # Match if first 8 chars are same, or one contains the other
            if addr['company_normalized']:
                if (cust_name[:8] == addr['company_normalized'][:8] or
                    cust_name in addr['company_normalized'] or
                    addr['company_normalized'] in cust_name):
                    return addr
            if addr['contact_normalized']:
                if (cust_name[:8] == addr['contact_normalized'][:8] or
                    cust_name in addr['contact_normalized'] or
                    addr['contact_normalized'] in cust_name):
                    return addr
    
    return None

def main():
    # File paths
    customers_file = 'attached_assets/customers_import_(1)_1767455004112.xlsx'
    csv_addressbook = 'attached_assets/export_1767454080709.csv'
    german_addressbook = 'attached_assets/Pasted-A-Nails-Nagelstudio-Ludwigstra-e-4-63739-Aschaffenburg-_1767454356355.txt'
    
    print("=" * 60)
    print("MERGING CUSTOMERS WITH ADDRESS BOOKS")
    print("=" * 60)
    
    # Load customers from Excel
    print("\n1. Loading customers from Excel...")
    wb_in = load_workbook(customers_file, data_only=True)
    ws_in = wb_in.active
    
    rows = list(ws_in.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]
    print(f"   Found {len(rows) - 1} customers")
    print(f"   Columns: {headers}")
    
    # Load address books
    print("\n2. Loading CSV address book...")
    csv_addresses = parse_csv_addressbook(csv_addressbook)
    print(f"   Found {len(csv_addresses)} addresses")
    
    print("\n3. Loading German TXT address book...")
    german_addresses = parse_german_addressbook(german_addressbook)
    print(f"   Found {len(german_addresses)} addresses")
    
    all_addresses = csv_addresses + german_addresses
    print(f"\n   Total addresses for lookup: {len(all_addresses)}")
    
    # Create output workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Customers with Addresses"
    
    # Define output headers (original + shipping address columns)
    output_headers = list(headers) + [
        'Shipping First Name', 'Shipping Last Name', 'Shipping Company',
        'Shipping Email', 'Shipping Phone', 'Shipping Street',
        'Shipping City', 'Shipping Zip Code', 'Shipping Country'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(output_headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Process each customer
    matched_count = 0
    no_match_count = 0
    
    print("\n4. Matching customers to addresses...")
    
    for row_idx, row in enumerate(rows[1:], 2):
        # Build customer dict
        customer = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                key = headers[i].lower().strip()
                customer[key] = str(value).strip() if value else ''
        
        # Get customer name and phone for matching
        customer['name'] = customer.get('name', '')
        customer['phone'] = customer.get('phone', customer.get('tel', ''))
        
        # Find matching address
        matched_addr = find_matching_address(customer, all_addresses)
        
        # Write original data
        for col, value in enumerate(row, 1):
            cell = ws_out.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
        
        # Write shipping address if found
        if matched_addr:
            matched_count += 1
            contact_parts = matched_addr.get('contact', '').split()
            first_name = contact_parts[0] if contact_parts else ''
            last_name = ' '.join(contact_parts[1:]) if len(contact_parts) > 1 else ''
            
            shipping_data = [
                first_name,
                last_name,
                matched_addr.get('company', ''),
                matched_addr.get('email', ''),
                matched_addr.get('phone', ''),
                matched_addr.get('street', ''),
                matched_addr.get('city', ''),
                matched_addr.get('zipCode', ''),
                matched_addr.get('country', ''),
            ]
            
            for col, value in enumerate(shipping_data, len(headers) + 1):
                cell = ws_out.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            no_match_count += 1
            for col in range(len(headers) + 1, len(output_headers) + 1):
                cell = ws_out.cell(row=row_idx, column=col, value='')
                cell.border = thin_border
    
    # Auto-adjust column widths
    for col in ws_out.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or '')) > max_length:
                    max_length = len(str(cell.value or ''))
            except:
                pass
        ws_out.column_dimensions[column].width = min(max_length + 2, 40)
    
    # Freeze header row
    ws_out.freeze_panes = 'A2'
    
    # Save output
    output_file = 'scripts/customers_merged_with_addresses.xlsx'
    wb_out.save(output_file)
    
    print(f"\n" + "=" * 60)
    print("RESULTS")
    print("=" * 60)
    print(f"Total customers: {len(rows) - 1}")
    print(f"Matched with address: {matched_count}")
    print(f"No match found: {no_match_count}")
    print(f"\nOutput saved to: {output_file}")
    print("=" * 60)

if __name__ == '__main__':
    main()
