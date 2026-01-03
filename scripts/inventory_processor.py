#!/usr/bin/env python3
"""
Inventory Processor Script
Generates SKUs matching Davie Supply's exact format and processes inventory data for import.

SKU Format: CATEGORY-PRODUCTPART (e.g., GP-SOGEPO for "SORAH Gel Polish 15ml" in "Gel Polish")
- Category: First letter of each word (multi-word) or first 3 chars (single word)
- Product: First 6 chars (1 word), first 3 of each (2 words), first 2 of first 3 (3+ words)
- Only adds -1, -2 suffix if duplicate exists

Usage:
    python scripts/inventory_processor.py input_file.tsv [supplier_file.tsv] output_file.xlsx
"""

import csv
import sys
import os
from typing import Set, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Vietnamese diacritics mapping (matching client/src/lib/vietnameseSearch.ts)
VIETNAMESE_MAP = {
    'á': 'a', 'à': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
    'ă': 'a', 'ắ': 'a', 'ằ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
    'â': 'a', 'ấ': 'a', 'ầ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
    'Á': 'A', 'À': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
    'Ă': 'A', 'Ắ': 'A', 'Ằ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
    'Â': 'A', 'Ấ': 'A', 'Ầ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
    'đ': 'd', 'Đ': 'D',
    'é': 'e', 'è': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
    'ê': 'e', 'ế': 'e', 'ề': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
    'É': 'E', 'È': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
    'Ê': 'E', 'Ế': 'E', 'Ề': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
    'í': 'i', 'ì': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
    'Í': 'I', 'Ì': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
    'ó': 'o', 'ò': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
    'ô': 'o', 'ố': 'o', 'ồ': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
    'ơ': 'o', 'ớ': 'o', 'ờ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
    'Ó': 'O', 'Ò': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
    'Ô': 'O', 'Ố': 'O', 'Ồ': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
    'Ơ': 'O', 'Ớ': 'O', 'Ờ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
    'ú': 'u', 'ù': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
    'ư': 'u', 'ứ': 'u', 'ừ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
    'Ú': 'U', 'Ù': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
    'Ư': 'U', 'Ứ': 'U', 'Ừ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
    'ý': 'y', 'ỳ': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
    'Ý': 'Y', 'Ỳ': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y'
}


def normalize_for_sku(text: str) -> str:
    """Normalize text for SKU generation - remove Vietnamese diacritics and non-alphanumeric chars."""
    import re
    normalized = ''.join(VIETNAMESE_MAP.get(c, c) for c in text)
    return re.sub(r'[^A-Za-z0-9]', '', normalized).upper()


def normalize_for_matching(text: str) -> str:
    """Normalize text for fuzzy matching - lowercase, remove diacritics, remove special chars."""
    import re
    normalized = ''.join(VIETNAMESE_MAP.get(c, c) for c in text)
    return re.sub(r'[^a-z0-9]', '', normalized.lower())


def get_category_part(category_name: str) -> str:
    """Generate category code from category name."""
    words = [w for w in category_name.split() if w and w not in ['&', 'and', '-', '/']]
    if len(words) > 1:
        category_part = ''.join(normalize_for_sku(w)[0] if normalize_for_sku(w) else '' for w in words)
    else:
        category_part = normalize_for_sku(category_name)[:3]
    return category_part.upper() if category_part else 'GEN'


def get_product_part(product_name: str) -> str:
    """Generate product code from product name."""
    words = [w for w in product_name.split() if w]
    if not words:
        return 'ITEM'
    if len(words) == 1:
        product_part = normalize_for_sku(words[0])[:6]
    elif len(words) == 2:
        product_part = normalize_for_sku(words[0])[:3] + normalize_for_sku(words[1])[:3]
    else:
        product_part = ''.join(normalize_for_sku(w)[:2] for w in words[:3])
    return product_part.upper() if product_part else 'ITEM'


def generate_sku(category: str, product_name: str, existing_skus: Set[str]) -> str:
    """Generate a unique SKU matching Davie Supply's exact format."""
    cat_part = get_category_part(category)
    prod_part = get_product_part(product_name)
    base_sku = f"{cat_part}-{prod_part}"
    if base_sku.upper() not in existing_skus:
        existing_skus.add(base_sku.upper())
        return base_sku
    counter = 1
    while f"{base_sku}-{counter}".upper() in existing_skus:
        counter += 1
    final_sku = f"{base_sku}-{counter}"
    existing_skus.add(final_sku.upper())
    return final_sku


def clean_price(value: str) -> float:
    """Clean price value, handling commas and #N/A."""
    if not value or str(value).strip() in ['#N/A', 'Loading...', '', 'None']:
        return 0.0
    cleaned = str(value).replace(',', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def load_supplier_mapping(supplier_file: str) -> Dict[str, str]:
    """
    Load supplier data and create a mapping from product name to supplier.
    Uses fuzzy matching to handle variations in product names.
    Always uses the newest supplier (by Stock Date) when there are duplicates.
    """
    from datetime import datetime
    
    # Store supplier with date for comparison: {key: (supplier, date)}
    supplier_with_dates: Dict[str, tuple] = {}
    
    delimiter = '\t' if supplier_file.endswith('.tsv') or supplier_file.endswith('.txt') else ','
    
    print(f"📦 Loading supplier data from: {supplier_file}")
    
    with open(supplier_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        
        for row in reader:
            product_name = row.get('Product name', '').strip()
            supplier = row.get('Supplier', '').strip()
            stock_date_str = row.get('Stock Date', '').strip()
            
            if not product_name or not supplier:
                continue
            
            # Parse date for comparison (format: YYYY-MM-DD)
            try:
                stock_date = datetime.strptime(stock_date_str, '%Y-%m-%d') if stock_date_str else datetime.min
            except ValueError:
                stock_date = datetime.min
            
            # Normalized name for matching
            normalized_name = normalize_for_matching(product_name)
            
            # Only update if this is newer than existing entry
            if normalized_name not in supplier_with_dates or stock_date > supplier_with_dates[normalized_name][1]:
                supplier_with_dates[normalized_name] = (supplier, stock_date)
            
            # Also store by SKU if available
            sku = row.get('SKU', '').strip()
            if sku:
                sku_key = f"sku:{sku.upper()}"
                if sku_key not in supplier_with_dates or stock_date > supplier_with_dates[sku_key][1]:
                    supplier_with_dates[sku_key] = (supplier, stock_date)
    
    # Extract just the supplier names (drop dates)
    supplier_map = {key: value[0] for key, value in supplier_with_dates.items()}
    
    print(f"   ✅ Loaded {len(supplier_map)} supplier mappings (using newest dates)")
    return supplier_map


def find_supplier(product_name: str, reference: str, supplier_map: Dict[str, str]) -> str:
    """
    Find supplier for a product using fuzzy matching.
    Tries multiple matching strategies.
    """
    # Strategy 1: Try exact reference/SKU match
    if reference:
        sku_key = f"sku:{reference.upper()}"
        if sku_key in supplier_map:
            return supplier_map[sku_key]
    
    # Strategy 2: Try normalized product name match
    normalized_name = normalize_for_matching(product_name)
    if normalized_name in supplier_map:
        return supplier_map[normalized_name]
    
    # Strategy 3: Try partial matching (contains)
    for key, supplier in supplier_map.items():
        if not key.startswith('sku:'):
            # Check if the supplier product name is contained in our product name
            if key in normalized_name or normalized_name in key:
                return supplier
    
    # Strategy 4: Try matching first significant words
    product_words = normalized_name[:15]  # First 15 chars
    for key, supplier in supplier_map.items():
        if not key.startswith('sku:') and key.startswith(product_words):
            return supplier
    
    return ''  # No match found


def process_inventory(input_file: str, supplier_file: str, output_file: str):
    """
    Process inventory file and generate new SKUs.
    Outputs to Excel format matching the import template.
    """
    products = []
    existing_skus: Set[str] = set()
    
    # Load supplier mapping if file provided
    supplier_map = {}
    if supplier_file and os.path.exists(supplier_file):
        supplier_map = load_supplier_mapping(supplier_file)
    
    delimiter = '\t' if input_file.endswith('.tsv') or input_file.endswith('.txt') else ','
    
    print(f"📂 Reading inventory from: {input_file}")
    
    matched_suppliers = 0
    
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        
        for row in reader:
            product_name = row.get('Product name', '').strip()
            category = row.get('Category', '').strip()
            reference = row.get('Reference', '').strip()
            
            if not product_name:
                continue
            
            new_sku = generate_sku(category, product_name, existing_skus)
            
            # Find supplier using reference or product name matching
            supplier = find_supplier(product_name, reference, supplier_map)
            if supplier:
                matched_suppliers += 1
            
            products.append({
                'Name': product_name,
                'Vietnamese Name': '',
                'SKU': new_sku,
                'Barcode': '',
                'Category': category,
                'Supplier': supplier,
                'Warehouse': '',
                'Warehouse Location': '',
                'Quantity': 0,
                'Low Stock Alert': 10,
                'Price CZK': clean_price(row.get('Price CZK', '0')),
                'Price EUR': clean_price(row.get('Price EUR', '0')),
                'Price USD': 0.0,
                'Wholesale Price CZK': 0.0,
                'Wholesale Price EUR': 0.0,
                'Import Cost USD': 0.0,
                'Import Cost EUR': clean_price(row.get('Imp. Cost EUR', '0')),
                'Import Cost CZK': clean_price(row.get('Imp. Cost CZK', '0')),
                'Weight (kg)': 0.0,
                'Length (cm)': 0.0,
                'Width (cm)': 0.0,
                'Height (cm)': 0.0,
                'Description': '',
                'Shipment Notes': '',
            })
    
    print(f"✅ Processed {len(products)} products")
    print(f"   🏭 Matched suppliers for {matched_suppliers}/{len(products)} products ({100*matched_suppliers//len(products)}%)")
    
    # Count unique vs duplicated SKUs
    dup_count = sum(1 for p in products if p['SKU'].rsplit('-', 1)[-1].isdigit() and '-' in p['SKU'])
    print(f"   📊 {len(products) - dup_count} unique base SKUs, {dup_count} with suffix counters")
    
    # Output to Excel
    if output_file.endswith('.xlsx') and HAS_OPENPYXL:
        print(f"💾 Writing Excel output to: {output_file}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        headers = [
            'Name', 'Vietnamese Name', 'SKU', 'Barcode', 'Category', 'Supplier',
            'Warehouse', 'Warehouse Location', 'Quantity', 'Low Stock Alert',
            'Price CZK', 'Price EUR', 'Price USD', 'Wholesale Price CZK', 'Wholesale Price EUR',
            'Import Cost USD', 'Import Cost EUR', 'Import Cost CZK',
            'Weight (kg)', 'Length (cm)', 'Width (cm)', 'Height (cm)',
            'Description', 'Shipment Notes'
        ]
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, product in enumerate(products, 2):
            for col_idx, header in enumerate(headers, 1):
                value = product.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_file)
    else:
        print(f"💾 Writing CSV output to: {output_file}")
        
        headers = [
            'Name', 'Vietnamese Name', 'SKU', 'Barcode', 'Category', 'Supplier',
            'Warehouse', 'Warehouse Location', 'Quantity', 'Low Stock Alert',
            'Price CZK', 'Price EUR', 'Price USD', 'Wholesale Price CZK', 'Wholesale Price EUR',
            'Import Cost USD', 'Import Cost EUR', 'Import Cost CZK',
            'Weight (kg)', 'Length (cm)', 'Width (cm)', 'Height (cm)',
            'Description', 'Shipment Notes'
        ]
        
        with open(output_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(products)
    
    print(f"🚀 Done! Generated {len(products)} products ready for import")
    
    # Show sample with suppliers
    print("\n📋 Sample output (first 15 products with suppliers):")
    print("-" * 100)
    products_with_supplier = [p for p in products if p['Supplier']]
    for product in products_with_supplier[:15]:
        print(f"  {product['SKU']:18} | {product['Supplier'][:30]:30} | {product['Name'][:35]}")
    
    if not products_with_supplier:
        print("  (No supplier matches found)")


def main():
    # Default files
    input_file = 'attached_assets/Pasted-Product-name-Reference-Category-SKU-Price-CZK-Price-EUR_1767405059741.txt'
    supplier_file = 'attached_assets/Pasted--Product-name-Stock-Date-Supplier-SKU-Quantity-Cost-Pri_1767406214298.txt'
    output_file = 'inventory_import_ready.xlsx'
    
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        supplier_file = sys.argv[2]
    if len(sys.argv) >= 4:
        output_file = sys.argv[3]
    
    if not os.path.exists(input_file):
        print(f"❌ Error: Input file not found: {input_file}")
        sys.exit(1)
    
    if supplier_file and not os.path.exists(supplier_file):
        print(f"⚠️ Warning: Supplier file not found: {supplier_file}")
        supplier_file = None
    
    process_inventory(input_file, supplier_file, output_file)


if __name__ == '__main__':
    main()
