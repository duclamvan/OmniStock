import os
import sys
import time
import re
import json
import openpyxl
import requests

# Configuration
INPUT_FILE = 'attached_assets/customers_comprehensive_1768685976098.xlsx'
OUTPUT_FILE = 'attached_assets/customers_comprehensive_with_fb_names.xlsx'
FACEBOOK_URL_COL = 10  # Column K (0-indexed)
FACEBOOK_NAME_COL = 9  # Column J (0-indexed)
FACEBOOK_NUMERIC_ID_COL = 11  # Column L (0-indexed) - New column for numeric ID
BATCH_SIZE = 10  # Process this many URLs per API call (Bright Data handles batches)
SAVE_EVERY = 50  # Save progress every N rows

# Bright Data Facebook Profiles - Collect by URL API
BRIGHT_DATA_DATASET_ID = 'gd_mf0urb782734ik94dz'

def extract_username_from_url(url):
    """Extract Facebook username/ID from URL"""
    if not url:
        return None
    
    clean_url = url.strip()
    
    # Pattern 1: profile.php?id=NUMERIC_ID
    match = re.search(r'facebook\.com\/profile\.php\?id=(\d+)', clean_url, re.I)
    if match:
        return match.group(1)
    
    # Pattern 2: /people/Name/NUMERIC_ID
    match = re.search(r'facebook\.com\/people\/[^\/]+\/(\d+)', clean_url, re.I)
    if match:
        return match.group(1)
    
    # Pattern 3: Any numeric ID (10+ digits)
    match = re.search(r'facebook\.com\/.*?(\d{10,})', clean_url, re.I)
    if match:
        return match.group(1)
    
    # Pattern 4: Standard username
    match = re.search(r'facebook\.com\/([a-zA-Z0-9._-]+)\/?(?:\?|$|#)?', clean_url, re.I)
    if match:
        username = match.group(1).lower()
        skip_paths = ['people', 'pages', 'groups', 'events', 'watch', 'marketplace', 'gaming', 'live', 'stories', 'reels', 'photo', 'photos', 'video', 'videos', 'share', 'sharer', 'pg']
        if username not in skip_paths:
            return match.group(1)
    
    return None

def build_fb_url(username):
    """Build Facebook URL from username/ID"""
    if username.isdigit():
        return f"https://www.facebook.com/profile.php?id={username}"
    return f"https://www.facebook.com/{username}"

def fetch_facebook_profiles_batch(api_key, batch):
    """Fetch multiple Facebook profiles using Bright Data API"""
    try:
        urls = [{"url": build_fb_url(username)} for _, _, username in batch]
        
        # Trigger Bright Data collection
        trigger_response = requests.post(
            f'https://api.brightdata.com/datasets/v3/trigger?dataset_id={BRIGHT_DATA_DATASET_ID}&include_errors=true&format=json',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            },
            json=urls
        )
        
        if trigger_response.status_code != 200:
            print(f"  Trigger error: {trigger_response.status_code} - {trigger_response.text}")
            return {}
        
        snapshot_info = trigger_response.json()
        snapshot_id = snapshot_info.get('snapshot_id')
        
        if not snapshot_id:
            print(f"  No snapshot ID returned: {snapshot_info}")
            return {}
        
        print(f"  Snapshot ID: {snapshot_id}, waiting for results...")
        
        # Poll for results (max 60 seconds)
        for attempt in range(60):
            time.sleep(1)
            
            status_response = requests.get(
                f'https://api.brightdata.com/datasets/v3/snapshot/{snapshot_id}?format=json',
                headers={'Authorization': f'Bearer {api_key}'}
            )
            
            if status_response.status_code == 200:
                results = status_response.json()
                if results:
                    # Map results back to usernames
                    results_map = {}
                    for item in results:
                        name = item.get('name') or item.get('account') or item.get('handle')
                        numeric_id = item.get('fbid') or item.get('id') or item.get('user_id')
                        url = item.get('url') or ''
                        
                        # Extract username from result URL to match back
                        username = extract_username_from_url(url)
                        if username and (name or numeric_id):
                            results_map[username.lower()] = {
                                'name': name,
                                'numeric_id': str(numeric_id) if numeric_id else None
                            }
                    
                    return results_map
            elif status_response.status_code == 202:
                # Still processing
                if attempt % 10 == 0:
                    print(f"  Still processing... ({attempt}s)")
                continue
            else:
                print(f"  Status check error: {status_response.status_code}")
                break
        
        print("  Timeout waiting for results")
        return {}
        
    except Exception as e:
        print(f"  Batch error: {e}")
        return {}

def main():
    api_key = os.environ.get('BRIGHT_DATA_API_KEY')
    if not api_key:
        print("ERROR: BRIGHT_DATA_API_KEY not found in environment")
        sys.exit(1)
    
    print(f"Loading workbook: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Collect rows needing processing
    rows_to_process = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fb_url = row[FACEBOOK_URL_COL].value
        fb_name = row[FACEBOOK_NAME_COL].value
        if fb_url and str(fb_url).strip() and (not fb_name or not str(fb_name).strip()):
            username = extract_username_from_url(str(fb_url))
            if username:
                rows_to_process.append((row_idx, fb_url, username))
    
    total = len(rows_to_process)
    print(f"Found {total} rows to process")
    print(f"Processing in batches of {BATCH_SIZE} using Bright Data API...")
    
    success_count = 0
    error_count = 0
    
    # Process in batches
    for batch_start in range(0, total, BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, total)
        batch = rows_to_process[batch_start:batch_end]
        
        print(f"\n[Batch {batch_start//BATCH_SIZE + 1}/{(total + BATCH_SIZE - 1)//BATCH_SIZE}] Processing rows {batch_start+1}-{batch_end}...")
        
        results = fetch_facebook_profiles_batch(api_key, batch)
        
        # Apply results to worksheet
        for row_idx, fb_url, username in batch:
            result = results.get(username.lower())
            if result:
                name = result.get('name')
                numeric_id = result.get('numeric_id')
                
                if name:
                    ws.cell(row=row_idx, column=FACEBOOK_NAME_COL + 1).value = name
                    print(f"  Row {row_idx}: {name}")
                    success_count += 1
                else:
                    print(f"  Row {row_idx}: No name for {username}")
                    error_count += 1
                
                if numeric_id:
                    ws.cell(row=row_idx, column=FACEBOOK_NUMERIC_ID_COL + 1).value = numeric_id
            else:
                print(f"  Row {row_idx}: No result for {username}")
                error_count += 1
        
        # Save progress periodically
        if batch_end % SAVE_EVERY == 0 or batch_end == total:
            print(f"  Saving progress ({batch_end} rows processed)...")
            wb.save(OUTPUT_FILE)
    
    # Final save
    wb.save(OUTPUT_FILE)
    print(f"\n=== COMPLETE ===")
    print(f"Output saved to: {OUTPUT_FILE}")
    print(f"Success: {success_count}, Errors: {error_count}")

if __name__ == "__main__":
    main()
