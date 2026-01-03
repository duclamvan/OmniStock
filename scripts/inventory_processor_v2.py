#!/usr/bin/env python3
"""
Inventory Processor Script V2
Processes NS-Management inventory export with supplier and import costs.

Input format (CSV from NS-Management):
- Product name, Reference, Category, SKU, Price CZK, Price EUR, 
  Imp. Cost USD, Imp. Cost EUR, Imp. Cost CZK, Supplier, Weight (g), 
  Imported (pcs), Stock (pcs), Total orders

Output format (Excel for Davie Supply import):
- Name, Vietnamese Name, SKU, Barcode, Category, Supplier, etc.

Usage:
    python scripts/inventory_processor_v2.py input_file.csv [output_file.xlsx]
"""

import csv
import sys
import os
from typing import Set, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

VIETNAMESE_MAP = {
    'á': 'a', 'à': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
    'ă': 'a', 'ắ': 'a', 'ằ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
    'â': 'a', 'ấ': 'a', 'ầ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
    'Á': 'A', 'À': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
    'Ă': 'A', 'Ắ': 'A', 'Ằ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
    'Â': 'A', 'Ấ': 'A', 'Ầ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
    'đ': 'd', 'Đ': 'D',
    'é': 'e', 'è': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
    'ê': 'e', 'ế': 'e', 'ề': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
    'É': 'E', 'È': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
    'Ê': 'E', 'Ế': 'E', 'Ề': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
    'í': 'i', 'ì': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
    'Í': 'I', 'Ì': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
    'ó': 'o', 'ò': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
    'ô': 'o', 'ố': 'o', 'ồ': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
    'ơ': 'o', 'ớ': 'o', 'ờ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
    'Ó': 'O', 'Ò': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
    'Ô': 'O', 'Ố': 'O', 'Ồ': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
    'Ơ': 'O', 'Ớ': 'O', 'Ờ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
    'ú': 'u', 'ù': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
    'ư': 'u', 'ứ': 'u', 'ừ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
    'Ú': 'U', 'Ù': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
    'Ư': 'U', 'Ứ': 'U', 'Ừ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
    'ý': 'y', 'ỳ': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
    'Ý': 'Y', 'Ỳ': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y'
}


def normalize_for_sku(text: str) -> str:
    """Normalize text for SKU generation - remove Vietnamese diacritics and non-alphanumeric chars."""
    import re
    normalized = ''.join(VIETNAMESE_MAP.get(c, c) for c in text)
    return re.sub(r'[^A-Za-z0-9]', '', normalized).upper()


def get_category_part(category_name: str) -> str:
    """Generate category code from category name."""
    words = [w for w in category_name.split() if w and w not in ['&', 'and', '-', '/']]
    if len(words) > 1:
        category_part = ''.join(normalize_for_sku(w)[0] if normalize_for_sku(w) else '' for w in words)
    else:
        category_part = normalize_for_sku(category_name)[:3]
    return category_part.upper() if category_part else 'GEN'


def get_product_part(product_name: str) -> str:
    """Generate product code from product name."""
    words = [w for w in product_name.split() if w]
    if not words:
        return 'ITEM'
    if len(words) == 1:
        product_part = normalize_for_sku(words[0])[:6]
    elif len(words) == 2:
        product_part = normalize_for_sku(words[0])[:3] + normalize_for_sku(words[1])[:3]
    else:
        product_part = ''.join(normalize_for_sku(w)[:2] for w in words[:3])
    return product_part.upper() if product_part else 'ITEM'


def generate_sku(category: str, product_name: str, existing_skus: Set[str]) -> str:
    """Generate a unique SKU matching Davie Supply's exact format."""
    cat_part = get_category_part(category)
    prod_part = get_product_part(product_name)
    base_sku = f"{cat_part}-{prod_part}"
    if base_sku.upper() not in existing_skus:
        existing_skus.add(base_sku.upper())
        return base_sku
    counter = 1
    while f"{base_sku}-{counter}".upper() in existing_skus:
        counter += 1
    final_sku = f"{base_sku}-{counter}"
    existing_skus.add(final_sku.upper())
    return final_sku


def clean_price(value: str) -> float:
    """Clean price value, handling commas and #N/A."""
    if not value or str(value).strip() in ['#N/A', 'Loading...', '', 'None', 'N/A']:
        return 0.0
    cleaned = str(value).replace(',', '').replace('"', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def clean_weight_to_kg(weight_g: str) -> float:
    """Convert weight from grams to kg."""
    weight = clean_price(weight_g)
    return weight / 1000.0 if weight > 0 else 0.0


def process_inventory(input_file: str, output_file: str):
    """
    Process NS-Management inventory export and generate import file.
    Uses existing SKUs from the file.
    """
    products = []
    existing_skus: Set[str] = set()
    
    delimiter = '\t' if input_file.endswith('.tsv') or input_file.endswith('.txt') else ','
    
    print(f"📂 Reading inventory from: {input_file}")
    
    with_supplier = 0
    with_cost_usd = 0
    
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        
        # Print available columns for debugging
        print(f"   📋 Columns found: {reader.fieldnames}")
        
        for row in reader:
            product_name = row.get('Product name', '').strip()
            category = row.get('Category', '').strip()
            existing_sku = row.get('SKU', '').strip()
            reference = row.get('Reference', '').strip()
            supplier = row.get('Supplier', '').strip()
            
            if not product_name:
                continue
            
            # Use existing SKU if available, otherwise generate new one
            if existing_sku and existing_sku.upper() not in existing_skus:
                sku = existing_sku
                existing_skus.add(existing_sku.upper())
            else:
                sku = generate_sku(category, product_name, existing_skus)
            
            if supplier:
                with_supplier += 1
            
            import_cost_usd = clean_price(row.get('Imp. Cost USD', '0'))
            if import_cost_usd > 0:
                with_cost_usd += 1
            
            products.append({
                'Name': product_name,
                'Vietnamese Name': '',
                'SKU': sku,
                'Barcode': '',
                'Category': category,
                'Supplier': supplier,
                'Warehouse': '',
                'Warehouse Location': '',
                'Quantity': 0,
                'Low Stock Alert': 10,
                'Price CZK': clean_price(row.get('Price CZK', '0')),
                'Price EUR': clean_price(row.get('Price EUR', '0')),
                'Price USD': 0.0,
                'Wholesale Price CZK': 0.0,
                'Wholesale Price EUR': 0.0,
                'Import Cost USD': import_cost_usd,
                'Import Cost EUR': clean_price(row.get('Imp. Cost EUR', '0')),
                'Import Cost CZK': clean_price(row.get('Imp. Cost CZK', '0')),
                'Weight (kg)': clean_weight_to_kg(row.get('Weight (g)', '0')),
                'Length (cm)': 0.0,
                'Width (cm)': 0.0,
                'Height (cm)': 0.0,
                'Description': '',
                'Shipment Notes': reference,
            })
    
    print(f"✅ Processed {len(products)} products")
    print(f"   🏭 Products with supplier: {with_supplier}/{len(products)} ({100*with_supplier//max(len(products),1)}%)")
    print(f"   💵 Products with USD import cost: {with_cost_usd}/{len(products)} ({100*with_cost_usd//max(len(products),1)}%)")
    
    # Output to Excel
    if output_file.endswith('.xlsx') and HAS_OPENPYXL:
        print(f"💾 Writing Excel output to: {output_file}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        headers = [
            'Name', 'Vietnamese Name', 'SKU', 'Barcode', 'Category', 'Supplier',
            'Warehouse', 'Warehouse Location', 'Quantity', 'Low Stock Alert',
            'Price CZK', 'Price EUR', 'Price USD', 'Wholesale Price CZK', 'Wholesale Price EUR',
            'Import Cost USD', 'Import Cost EUR', 'Import Cost CZK',
            'Weight (kg)', 'Length (cm)', 'Width (cm)', 'Height (cm)',
            'Description', 'Shipment Notes'
        ]
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_idx, product in enumerate(products, 2):
            for col_idx, header in enumerate(headers, 1):
                value = product.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        wb.save(output_file)
    else:
        print(f"💾 Writing CSV output to: {output_file}")
        
        headers = [
            'Name', 'Vietnamese Name', 'SKU', 'Barcode', 'Category', 'Supplier',
            'Warehouse', 'Warehouse Location', 'Quantity', 'Low Stock Alert',
            'Price CZK', 'Price EUR', 'Price USD', 'Wholesale Price CZK', 'Wholesale Price EUR',
            'Import Cost USD', 'Import Cost EUR', 'Import Cost CZK',
            'Weight (kg)', 'Length (cm)', 'Width (cm)', 'Height (cm)',
            'Description', 'Shipment Notes'
        ]
        
        with open(output_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(products)
    
    print(f"🚀 Done! Generated {len(products)} products ready for import")
    
    # Show sample with suppliers and costs
    print("\n📋 Sample output (first 20 products with suppliers):")
    print("-" * 120)
    print(f"{'SKU':15} | {'Supplier':35} | {'Cost USD':>10} | {'Cost EUR':>10} | {'Name':30}")
    print("-" * 120)
    
    for product in products[:20]:
        if product['Supplier'] or product['Import Cost USD'] > 0:
            supplier = (product['Supplier'][:32] + '...') if len(product['Supplier']) > 35 else product['Supplier']
            name = (product['Name'][:27] + '...') if len(product['Name']) > 30 else product['Name']
            cost_usd = f"${product['Import Cost USD']:.2f}" if product['Import Cost USD'] > 0 else '-'
            cost_eur = f"€{product['Import Cost EUR']:.2f}" if product['Import Cost EUR'] > 0 else '-'
            print(f"{product['SKU']:15} | {supplier:35} | {cost_usd:>10} | {cost_eur:>10} | {name}")


def main():
    input_file = 'attached_assets/NS-Management_-_INVENTORY_(2)_1767482264908.csv'
    output_file = 'inventory_import_ready.xlsx'
    
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    
    if not os.path.exists(input_file):
        print(f"❌ Error: Input file not found: {input_file}")
        sys.exit(1)
    
    process_inventory(input_file, output_file)


if __name__ == '__main__':
    main()
