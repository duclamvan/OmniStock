#!/usr/bin/env python3
"""
Create XLSX file with merged customer and address data for import.
"""

import json
import psycopg2
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    # Connect to database
    db_url = os.environ.get('DATABASE_URL')
    conn = psycopg2.connect(db_url)
    cur = conn.cursor()
    
    # Get all customers with their shipping addresses
    cur.execute("""
        SELECT 
            c.id, c.name, c.email, c.phone, c.address, c.city, c.zip_code, c.country,
            c.facebook_id, c.facebook_name, c.facebook_url,
            c.ico, c.dic, c.vat_number,
            c.preferred_currency, c.preferred_language, c.notes, c.type,
            c.billing_first_name, c.billing_last_name, c.billing_company,
            c.billing_email, c.billing_tel, c.billing_street, c.billing_street_number,
            c.billing_city, c.billing_zip_code, c.billing_country,
            sa.first_name as ship_first_name, sa.last_name as ship_last_name,
            sa.company as ship_company, sa.email as ship_email, sa.tel as ship_tel,
            sa.street as ship_street, sa.street_number as ship_street_number,
            sa.city as ship_city, sa.zip_code as ship_zip_code, sa.country as ship_country,
            sa.label as ship_label
        FROM customers c
        LEFT JOIN customer_shipping_addresses sa ON c.id = sa.customer_id
        ORDER BY c.name, sa.is_primary DESC, sa.created_at
    """)
    
    rows = cur.fetchall()
    print(f"Found {len(rows)} customer-address combinations")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers with Addresses"
    
    # Headers
    headers = [
        'Name', 'Email', 'Phone', 'Address', 'City', 'Zip Code', 'Country', 'Type',
        'Facebook ID', 'Facebook Name', 'Facebook URL',
        'ICO', 'DIC', 'VAT Number',
        'Preferred Currency', 'Preferred Language', 'Notes',
        'Billing First Name', 'Billing Last Name', 'Billing Company',
        'Billing Email', 'Billing Phone', 'Billing Street', 'Billing Street Number',
        'Billing City', 'Billing Zip Code', 'Billing Country',
        'Shipping First Name', 'Shipping Last Name', 'Shipping Company',
        'Shipping Email', 'Shipping Phone', 'Shipping Street', 'Shipping Street Number',
        'Shipping City', 'Shipping Zip Code', 'Shipping Country', 'Shipping Label'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, row in enumerate(rows, 2):
        data = [
            row[1],   # name
            row[2],   # email
            row[3],   # phone
            row[4],   # address
            row[5],   # city
            row[6],   # zip_code
            row[7],   # country
            row[17],  # type
            row[8],   # facebook_id
            row[9],   # facebook_name
            row[10],  # facebook_url
            row[11],  # ico
            row[12],  # dic
            row[13],  # vat_number
            row[14],  # preferred_currency
            row[15],  # preferred_language
            row[16],  # notes
            row[18],  # billing_first_name
            row[19],  # billing_last_name
            row[20],  # billing_company
            row[21],  # billing_email
            row[22],  # billing_tel
            row[23],  # billing_street
            row[24],  # billing_street_number
            row[25],  # billing_city
            row[26],  # billing_zip_code
            row[27],  # billing_country
            row[28],  # ship_first_name
            row[29],  # ship_last_name
            row[30],  # ship_company
            row[31],  # ship_email
            row[32],  # ship_tel
            row[33],  # ship_street
            row[34],  # ship_street_number
            row[35],  # ship_city
            row[36],  # ship_zip_code
            row[37],  # ship_country
            row[38],  # ship_label
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save
    output_file = 'scripts/customers_with_addresses_export.xlsx'
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")
    
    # Stats
    cur.execute("SELECT COUNT(DISTINCT id) FROM customers")
    total_customers = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM customer_shipping_addresses")
    total_addresses = cur.fetchone()[0]
    
    print(f"Total customers: {total_customers}")
    print(f"Total shipping addresses: {total_addresses}")
    print(f"Total rows in export: {len(rows)}")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
