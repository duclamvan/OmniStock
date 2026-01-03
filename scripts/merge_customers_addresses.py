#!/usr/bin/env python3
"""
Merge customers from Excel with addresses from both addressbooks.
Matches by phone number, company name, or contact name.
"""

import json
import re
import sys
from pathlib import Path

# Try to import openpyxl for reading Excel files
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

def normalize_phone(phone: str) -> str:
    """Normalize phone number for comparison (last 9 digits)."""
    if not phone:
        return ""
    digits = re.sub(r'[^\d]', '', phone)
    return digits[-9:] if len(digits) >= 9 else digits

def normalize_name(name: str) -> str:
    """Normalize name for fuzzy matching."""
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', name.lower())

def parse_csv_addressbook(filepath: str) -> list:
    """Parse the semicolon-separated CSV address book."""
    addresses = []
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        lines = f.readlines()
    
    if not lines:
        return addresses
    
    headers = [h.strip() for h in lines[0].split(';')]
    
    for line in lines[1:]:
        if not line.strip():
            continue
        values = line.split(';')
        row = dict(zip(headers, values))
        
        address = {
            'source': 'csv',
            'company': row.get('NAZEV', '').strip(),
            'street': row.get('ULICE_CP', '').strip(),
            'city': row.get('MESTO', '').strip(),
            'zipCode': row.get('PSC', '').strip(),
            'country': row.get('ZEME', 'CZ').strip(),
            'email': row.get('EMAIL', '').strip(),
            'contact': row.get('KONTAKTNI_OS', '').strip(),
            'phone': row.get('TELEFON', '').strip(),
        }
        
        # Convert country codes
        country_map = {'CZ': 'Czech Republic', 'DE': 'Germany', 'AT': 'Austria', 
                      'SK': 'Slovakia', 'PL': 'Poland', 'DK': 'Denmark', 'BE': 'Belgium'}
        address['country'] = country_map.get(address['country'], address['country'])
        
        if address['street'] or address['company']:
            addresses.append(address)
    
    return addresses

def parse_german_addressbook(filepath: str) -> list:
    """Parse the German TXT format address book."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    entries = re.split(r'\nBearbeiten\s*\n+Löschen\s*\n?', content)
    addresses = []
    
    for entry in entries:
        lines = [l.strip() for l in entry.strip().split('\n') if l.strip()]
        if len(lines) < 3:
            continue
        
        address = {
            'source': 'german_txt',
            'company': '',
            'contact': '',
            'street': '',
            'city': '',
            'zipCode': '',
            'country': 'Germany',
            'email': '',
            'phone': '',
        }
        
        # Find email
        for line in lines:
            if '@' in line:
                address['email'] = line
                break
        
        # Find postal code + city
        postal_idx = None
        for i, line in enumerate(lines):
            match = re.match(r'^(\d{4,5})\s+(.+?),\s*(\w{2})$', line)
            if match:
                address['zipCode'] = match.group(1)
                address['city'] = match.group(2)
                country_map = {'DE': 'Germany', 'AT': 'Austria', 'NL': 'Netherlands',
                              'DK': 'Denmark', 'CH': 'Switzerland', 'BE': 'Belgium', 'FR': 'France'}
                address['country'] = country_map.get(match.group(3).upper(), match.group(3))
                postal_idx = i
                break
        
        if postal_idx is None or postal_idx < 1:
            continue
        
        # Street is line before postal
        address['street'] = lines[postal_idx - 1]
        
        # Company/contact from first lines
        remaining = lines[:postal_idx - 1]
        if remaining:
            company_keywords = ['nails', 'beauty', 'studio', 'salon', 'center', 'shop', 'lounge', 'spa']
            first = remaining[0]
            if any(kw in first.lower() for kw in company_keywords):
                address['company'] = first
                if len(remaining) > 1 and not remaining[1].startswith('('):
                    address['contact'] = remaining[1]
            else:
                address['contact'] = first
                if len(remaining) > 1:
                    address['company'] = remaining[1]
        
        if address['street'] and address['zipCode']:
            addresses.append(address)
    
    return addresses

def parse_customers_excel(filepath: str) -> list:
    """Parse customers from Excel file."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    customers = []
    
    for row in rows[1:]:
        customer = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                customer[headers[i]] = str(value).strip() if value else ''
        
        if customer.get('name'):
            customers.append(customer)
    
    return customers

def match_address_to_customer(address: dict, customers: list) -> dict | None:
    """Try to match an address to a customer by phone, company, or contact name."""
    addr_phone = normalize_phone(address.get('phone', ''))
    addr_company = normalize_name(address.get('company', ''))
    addr_contact = normalize_name(address.get('contact', ''))
    
    for customer in customers:
        # Match by phone (most reliable)
        cust_phone = normalize_phone(customer.get('phone', ''))
        if addr_phone and cust_phone and addr_phone == cust_phone:
            return customer
        
        # Match by company name (partial)
        cust_name = normalize_name(customer.get('name', ''))
        if addr_company and cust_name:
            # Check if first 10 chars match or one contains the other
            if (addr_company[:10] == cust_name[:10] or 
                addr_company in cust_name or 
                cust_name in addr_company):
                return customer
        
        # Match by contact name
        if addr_contact and cust_name:
            if addr_contact == cust_name or addr_contact in cust_name or cust_name in addr_contact:
                return customer
    
    return None

def main():
    # File paths
    customers_file = 'attached_assets/customers_import_(1)_1767453838575.xlsx'
    csv_addressbook = 'attached_assets/export_1767454080709.csv'
    german_addressbook = 'attached_assets/Pasted-A-Nails-Nagelstudio-Ludwigstra-e-4-63739-Aschaffenburg-_1767454356355.txt'
    
    print("=" * 60)
    print("MERGING CUSTOMERS WITH ADDRESS BOOKS")
    print("=" * 60)
    
    # Parse all sources
    print("\n1. Parsing customer Excel file...")
    customers = parse_customers_excel(customers_file)
    print(f"   Found {len(customers)} customers")
    
    print("\n2. Parsing CSV address book...")
    csv_addresses = parse_csv_addressbook(csv_addressbook)
    print(f"   Found {len(csv_addresses)} addresses")
    
    print("\n3. Parsing German TXT address book...")
    german_addresses = parse_german_addressbook(german_addressbook)
    print(f"   Found {len(german_addresses)} addresses")
    
    # Combine all addresses
    all_addresses = csv_addresses + german_addresses
    print(f"\n   Total addresses to match: {len(all_addresses)}")
    
    # Match addresses to customers
    print("\n4. Matching addresses to customers...")
    matched = []
    unmatched = []
    
    for addr in all_addresses:
        customer = match_address_to_customer(addr, customers)
        if customer:
            matched.append({
                'customer_name': customer.get('name'),
                'customer_phone': customer.get('phone'),
                'address': addr,
            })
        else:
            unmatched.append(addr)
    
    print(f"   Matched: {len(matched)} addresses")
    print(f"   Unmatched: {len(unmatched)} addresses")
    
    # Create output for import
    print("\n5. Creating merged output...")
    
    # Build customer -> addresses map
    customer_addresses = {}
    for m in matched:
        name = m['customer_name']
        if name not in customer_addresses:
            customer_addresses[name] = {
                'customer': next((c for c in customers if c.get('name') == name), {}),
                'addresses': []
            }
        customer_addresses[name]['addresses'].append(m['address'])
    
    # Output summary
    print(f"\n   Customers with matched addresses: {len(customer_addresses)}")
    
    # Save results
    output = {
        'matched_customers': list(customer_addresses.values()),
        'unmatched_addresses': unmatched,
        'stats': {
            'total_customers': len(customers),
            'total_addresses': len(all_addresses),
            'matched': len(matched),
            'unmatched': len(unmatched),
        }
    }
    
    output_file = 'scripts/merged_customers_addresses.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\n   Saved to: {output_file}")
    
    # Show sample matches
    print("\n" + "=" * 60)
    print("SAMPLE MATCHES:")
    print("=" * 60)
    for name, data in list(customer_addresses.items())[:5]:
        print(f"\n   Customer: {name}")
        for addr in data['addresses'][:2]:
            print(f"      -> {addr.get('company', addr.get('contact', 'N/A'))}: {addr['street']}, {addr['city']}")
    
    print("\n" + "=" * 60)
    print("DONE!")
    print("=" * 60)

if __name__ == '__main__':
    main()
