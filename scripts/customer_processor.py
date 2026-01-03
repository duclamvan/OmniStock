#!/usr/bin/env python3
"""
Customer Processor Script
Processes customer data with address parsing via API.

Features:
- Converts User ID (replacing ° with .) to Facebook ID
- Batches address parsing via API with smart caching to reduce API costs
- Extracts names from addresses or generates placeholder names
- Outputs Excel format matching customer import template

Cost-saving strategies:
1. Caches parsed addresses to avoid duplicate API calls
2. Skips addresses marked as "no address in db", "Pickup", etc.
3. Batches requests with delay to avoid rate limiting

Usage:
    python scripts/customer_processor.py [input_file.txt] [output_file.xlsx]
"""

import csv
import sys
import os
import json
import time
import hashlib
import re
from typing import Dict, Optional, Tuple
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from urllib.parse import urlencode

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Cache file to store parsed addresses and avoid duplicate API calls
CACHE_FILE = 'scripts/.address_cache.json'

# Vietnamese diacritics mapping
VIETNAMESE_MAP = {
    'á': 'a', 'à': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
    'ă': 'a', 'ắ': 'a', 'ằ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
    'â': 'a', 'ấ': 'a', 'ầ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
    'Á': 'A', 'À': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
    'Ă': 'A', 'Ắ': 'A', 'Ằ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
    'Â': 'A', 'Ấ': 'A', 'Ầ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
    'đ': 'd', 'Đ': 'D',
    'é': 'e', 'è': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
    'ê': 'e', 'ế': 'e', 'ề': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
    'É': 'E', 'È': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
    'Ê': 'E', 'Ế': 'E', 'Ề': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
    'í': 'i', 'ì': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
    'Í': 'I', 'Ì': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
    'ó': 'o', 'ò': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
    'ô': 'o', 'ố': 'o', 'ồ': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
    'ơ': 'o', 'ớ': 'o', 'ờ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
    'Ó': 'O', 'Ò': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
    'Ô': 'O', 'Ố': 'O', 'Ồ': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
    'Ơ': 'O', 'Ớ': 'O', 'Ờ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
    'ú': 'u', 'ù': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
    'ư': 'u', 'ứ': 'u', 'ừ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
    'Ú': 'U', 'Ù': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
    'Ư': 'U', 'Ứ': 'U', 'Ừ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
    'ý': 'y', 'ỳ': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
    'Ý': 'Y', 'Ỳ': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y'
}

# Skip patterns - addresses that don't need API parsing
SKIP_PATTERNS = [
    'no address in db',
    'pickup',
    'no address',
    'n/a',
    'inspire nails',  # Just shop name, no address
]


def normalize_text(text: str) -> str:
    """Remove Vietnamese diacritics."""
    return ''.join(VIETNAMESE_MAP.get(c, c) for c in text)


def convert_user_id_to_facebook_id(user_id: str) -> str:
    """Convert User ID by replacing ° with . to create Facebook ID."""
    return user_id.replace('°', '.')


def should_skip_address(address: str) -> bool:
    """Check if address should be skipped (no real address data)."""
    if not address or len(address.strip()) < 5:
        return True
    lower = address.lower().strip()
    for pattern in SKIP_PATTERNS:
        if pattern in lower:
            return True
    return False


def get_address_hash(address: str) -> str:
    """Generate hash for address to use as cache key."""
    normalized = normalize_text(address.lower().strip())
    return hashlib.md5(normalized.encode('utf-8')).hexdigest()[:16]


def load_cache() -> Dict:
    """Load address cache from file."""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_cache(cache: Dict):
    """Save address cache to file."""
    try:
        os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"   ⚠️ Could not save cache: {e}")


def parse_address_basic(raw_address: str) -> Dict[str, str]:
    """
    Basic regex-based address parsing (fallback when API unavailable).
    Matches logic from server/routes.ts parseAddressBasic function.
    """
    fields = {
        'firstName': '',
        'lastName': '',
        'company': '',
        'email': '',
        'tel': '',
        'street': '',
        'streetNumber': '',
        'city': '',
        'zipCode': '',
        'country': ''
    }
    
    if not raw_address:
        return fields
    
    text = raw_address.strip()
    
    # Extract phone number
    phone_match = re.search(r'(?:tel[:\s]*|phone[:\s]*|dt[:\s]*)?(\+?[\d\s\-]{8,20})', text, re.IGNORECASE)
    if phone_match:
        fields['tel'] = re.sub(r'[\s\-]', '', phone_match.group(1))
    
    # Extract email
    email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', text)
    if email_match:
        fields['email'] = email_match.group(0)
    
    # Extract postal code (German/Czech format: 5 digits)
    zip_match = re.search(r'\b(\d{4,5})\b', text)
    if zip_match:
        fields['zipCode'] = zip_match.group(1)
    
    # Extract street with number (European format: Street Name Number)
    street_match = re.search(r'([A-Za-zäöüÄÖÜßáéíóúàèìòùâêîôûčďěňřšťůžÁÉÍÓÚÀÈÌÒÙÂÊÎÔÛČĎĚŇŘŠŤŮŽ\s\-\.]+)\s+(\d+[a-zA-Z]?(?:[-/]\d+)?)', text)
    if street_match:
        fields['street'] = street_match.group(1).strip()
        fields['streetNumber'] = street_match.group(2).strip()
    
    # Try to extract name (first part before semicolon often contains name)
    parts = [p.strip() for p in text.split(';') if p.strip()]
    if parts:
        # Look for Vietnamese name pattern or company name
        first_part = parts[0]
        
        # Check if it's a company name (contains "nail", "beauty", "studio", etc.)
        company_keywords = ['nail', 'beauty', 'studio', 'salon', 'spa', 'lounge']
        if any(kw in first_part.lower() for kw in company_keywords):
            fields['company'] = first_part
            # Look for person name in subsequent parts
            for part in parts[1:]:
                # Vietnamese names typically have 2-4 words
                words = part.split()
                if 2 <= len(words) <= 4 and not any(kw in part.lower() for kw in company_keywords):
                    name_words = part.split()
                    if len(name_words) >= 2:
                        fields['lastName'] = name_words[0]
                        fields['firstName'] = ' '.join(name_words[1:])
                    break
        else:
            # First part might be a name
            name_words = first_part.split()
            if 2 <= len(name_words) <= 4:
                fields['lastName'] = name_words[0]
                fields['firstName'] = ' '.join(name_words[1:])
    
    # Try to extract city (often after postal code or in semicolon-separated part)
    if fields['zipCode']:
        # Look for text after postal code
        city_match = re.search(rf'{fields["zipCode"]}\s+([A-Za-zäöüÄÖÜßáéíóúàèìòùâêîôûčďěňřšťůžÁÉÍÓÚÀÈÌÒÙÂÊÎÔÛČĎĚŇŘŠŤŮŽ\s\-]+)', text)
        if city_match:
            fields['city'] = city_match.group(1).strip().split(';')[0].strip()
    
    # Detect country from text
    country_patterns = {
        'Germany': ['germany', 'deutschland', 'germeny', 'de'],
        'Czech Republic': ['czech', 'česká', 'ceska', 'cz', 'praha', 'prague'],
        'Netherlands': ['nederland', 'netherlands', 'holland', 'nl'],
        'Austria': ['austria', 'österreich', 'osterreich', 'at'],
        'Spain': ['spain', 'españa', 'espana', 'es', 'bcn', 'barcelona']
    }
    
    lower_text = text.lower()
    for country, patterns in country_patterns.items():
        if any(p in lower_text for p in patterns):
            fields['country'] = country
            break
    
    # Default to Germany if no country detected (most addresses seem German)
    if not fields['country'] and (fields['street'] or fields['city']):
        fields['country'] = 'Germany'
    
    return fields


def parse_address_via_api(raw_address: str, api_url: str, session_cookie: str = None) -> Optional[Dict[str, str]]:
    """
    Parse address using the application's AI address parsing API.
    Returns None if API call fails.
    """
    try:
        data = json.dumps({'rawAddress': raw_address}).encode('utf-8')
        
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }
        if session_cookie:
            headers['Cookie'] = session_cookie
        
        req = Request(api_url, data=data, headers=headers, method='POST')
        
        with urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode('utf-8'))
            return result.get('fields', {})
    
    except HTTPError as e:
        if e.code == 401:
            print(f"   ⚠️ API requires authentication (401)")
        else:
            print(f"   ⚠️ API error {e.code}: {e.reason}")
        return None
    except URLError as e:
        print(f"   ⚠️ Connection error: {e.reason}")
        return None
    except Exception as e:
        print(f"   ⚠️ Parse error: {e}")
        return None


def extract_name_from_address(address: str) -> Tuple[str, str]:
    """
    Extract first and last name from address text.
    Returns (firstName, lastName) tuple.
    """
    if not address:
        return ('', '')
    
    # Split by semicolons
    parts = [p.strip() for p in address.split(';') if p.strip()]
    
    # Common phrases to skip (Vietnamese and German)
    skip_phrases = [
        'địa chỉ', 'dia chi', 'bạn gửi', 'ban gui', 'gửi về', 'gui ve',
        'd/c', 'address', 'ok e', 'đc', 'dc', 'của chị', 'cua chi',
        'em theo', 'anh gởi', 'cho nguyen', 'nhé', 'nhe', 'là:', 'la:',
        'gửi hàng', 'gui hang', 'chuyển cho', 'chuyen cho'
    ]
    
    # Shop/company indicators
    company_keywords = [
        'nail', 'beauty', 'studio', 'salon', 'spa', 'lounge', 'center', 
        'centrum', 'galerie', 'kaufland', 'rewe', 'arcaden', 'sushi',
        'kitchen', 'restaurant', 'shop', 'store', 'gbr', 'gmbh', 'nails'
    ]
    
    # Address indicators
    address_keywords = [
        'str', 'strasse', 'straße', 'gasse', 'platz', 'weg', 'allee',
        'ring', 'damm', 'chaussee', 'ufer', 'markt', 'bahnhof',
        'im ', 'in der', 'am ', 'an der', 'bei ', 'eingang'
    ]
    
    # Look for patterns like "Name Name" that look like person names
    name_pattern = re.compile(
        r'^([A-ZÄÖÜČĎĚŇŘŠŤŮŽ][a-zäöüáéíóúàèìòùâêîôûčďěňřšťůž]+)\s+'
        r'([A-ZÄÖÜČĎĚŇŘŠŤŮŽ][a-zäöüáéíóúàèìòùâêîôûčďěňřšťůž]+)'
        r'(?:\s+[A-ZÄÖÜČĎĚŇŘŠŤŮŽ][a-zäöüáéíóúàèìòùâêîôûčďěňřšťůž]+)?$',
        re.UNICODE
    )
    
    # Vietnamese name pattern (Nguyen Thi Van, Tran Duc Minh, etc.)
    vn_name_pattern = re.compile(
        r'^((?:Nguyen|Tran|Le|Pham|Hoang|Vu|Vo|Dang|Bui|Do|Ho|Ngo|Duong|Ly|'
        r'Nguyễn|Trần|Lê|Phạm|Hoàng|Vũ|Võ|Đặng|Bùi|Đỗ|Hồ|Ngô|Dương|Lý|'
        r'Huynh|Huỳnh|Phan|Truong|Trương|Dinh|Đinh|Luu|Lưu|Mai|Dao|Đào|'
        r'Ta|Tạ|Trinh|Trịnh|Quach|Quách|Cao|La|Lã|Khong|Không|Khuu|Khưu)'
        r')\s+(.+)$',
        re.IGNORECASE | re.UNICODE
    )
    
    for part in parts:
        lower_part = part.lower()
        
        # Skip common prefixes
        if any(phrase in lower_part for phrase in skip_phrases):
            continue
        
        # Skip if it looks like a company/shop name
        if any(kw in lower_part for kw in company_keywords):
            continue
        
        # Skip if it contains address indicators
        if any(kw in lower_part for kw in address_keywords):
            continue
        
        # Skip if mostly numbers (likely address or phone)
        if sum(c.isdigit() for c in part) > len(part) * 0.3:
            continue
        
        # Skip if too short or too long
        words = part.split()
        if len(words) < 2 or len(words) > 5:
            continue
        
        # Try Vietnamese name pattern
        vn_match = vn_name_pattern.match(part)
        if vn_match:
            return (vn_match.group(2), vn_match.group(1))
        
        # Try general name pattern
        name_match = name_pattern.match(part)
        if name_match:
            return (name_match.group(2), name_match.group(1))
        
        # Fallback: if 2-4 words and looks like a name
        if 2 <= len(words) <= 4:
            # Check if first word could be a family name
            first_word = words[0]
            if len(first_word) >= 2 and first_word[0].isupper():
                return (' '.join(words[1:]), words[0])
    
    # Second pass: look for any part with parentheses containing a name
    for part in parts:
        # Pattern like "Shop Name (Person Name)"
        paren_match = re.search(r'\(([^)]+)\)', part)
        if paren_match:
            inner = paren_match.group(1).strip()
            words = inner.split()
            if 2 <= len(words) <= 4:
                # Check if it looks like a name (not a location)
                lower_inner = inner.lower()
                if not any(kw in lower_inner for kw in company_keywords + address_keywords):
                    return (' '.join(words[1:]), words[0])
    
    return ('', '')


def process_customers(input_file: str, output_file: str, api_url: str = None, session_cookie: str = None):
    """
    Process customer file and generate import-ready Excel.
    """
    customers = []
    cache = load_cache()
    api_calls = 0
    cache_hits = 0
    skipped = 0
    
    print(f"📂 Reading customers from: {input_file}")
    print(f"   💾 Cache loaded with {len(cache)} entries")
    
    # Detect delimiter
    delimiter = '\t' if input_file.endswith('.tsv') or input_file.endswith('.txt') else ','
    
    with open(input_file, 'r', encoding='utf-8') as f:
        # Skip header if present
        first_line = f.readline()
        if 'User ID' in first_line and 'Address' in first_line:
            pass  # Header line, already consumed
        else:
            f.seek(0)  # Reset to beginning if no header
        
        reader = csv.reader(f, delimiter=delimiter)
        
        for row in reader:
            if not row or len(row) < 2:
                continue
            
            user_id = row[0].strip()
            raw_address = row[1].strip() if len(row) > 1 else ''
            
            # Skip header row
            if user_id.lower() == 'user id':
                continue
            
            # Convert User ID to Facebook ID (replace ° with .)
            facebook_id = convert_user_id_to_facebook_id(user_id)
            
            # Initialize customer record
            customer = {
                'Name': '',
                'Facebook Name': '',
                'Facebook ID': facebook_id,
                'Facebook URL': f'https://facebook.com/{facebook_id}' if facebook_id and not facebook_id.startswith('http') else '',
                'Email': '',
                'Phone': '',
                'Address': raw_address,
                'City': '',
                'Zip Code': '',
                'Country': '',
                'Notes': '',
                'Type': 'regular',
                'Preferred Language': 'vi',
                'Preferred Currency': 'EUR',
                # VAT/Tax fields
                'ICO': '',
                'DIC': '',
                'VAT ID': '',
                'VAT Number': '',
                'Tax ID': '',
                # Billing address fields
                'Billing First Name': '',
                'Billing Last Name': '',
                'Billing Company': '',
                'Billing Email': '',
                'Billing Phone': '',
                'Billing Street': '',
                'Billing Street Number': '',
                'Billing City': '',
                'Billing Zip Code': '',
                'Billing Country': '',
            }
            
            # Check if address should be skipped
            if should_skip_address(raw_address):
                # Try to extract name anyway
                first_name, last_name = extract_name_from_address(raw_address)
                if first_name or last_name:
                    customer['Name'] = f"{first_name} {last_name}".strip()
                else:
                    customer['Name'] = facebook_id  # Use Facebook ID as fallback name
                
                customer['Notes'] = 'No valid address'
                skipped += 1
                customers.append(customer)
                continue
            
            # Check cache first
            addr_hash = get_address_hash(raw_address)
            parsed = None
            
            if addr_hash in cache:
                parsed = cache[addr_hash]
                cache_hits += 1
            elif api_url:
                # Call API for parsing
                parsed = parse_address_via_api(raw_address, api_url, session_cookie)
                api_calls += 1
                
                if parsed:
                    cache[addr_hash] = parsed
                    # Small delay to avoid rate limiting
                    if api_calls % 10 == 0:
                        time.sleep(0.5)
                
                # Progress indicator
                if api_calls % 50 == 0:
                    print(f"   📡 API calls: {api_calls}, Cache hits: {cache_hits}")
            
            # Fallback to basic parsing if API failed or unavailable
            if not parsed:
                parsed = parse_address_basic(raw_address)
            
            # Populate customer record from parsed address
            first_name = parsed.get('firstName', '')
            last_name = parsed.get('lastName', '')
            
            # If no name from parsing, try to extract from raw address
            if not first_name and not last_name:
                first_name, last_name = extract_name_from_address(raw_address)
            
            # Build full name
            if first_name or last_name:
                customer['Name'] = f"{first_name} {last_name}".strip()
            else:
                customer['Name'] = facebook_id  # Fallback to Facebook ID
            
            # Contact info
            customer['Email'] = parsed.get('email', '')
            customer['Phone'] = parsed.get('tel', '')
            
            # Main address fields
            street = parsed.get('street', '')
            street_num = parsed.get('streetNumber', '')
            customer['Address'] = f"{street} {street_num}".strip() if street else raw_address
            customer['City'] = parsed.get('city', '')
            customer['Zip Code'] = parsed.get('zipCode', '')
            customer['Country'] = parsed.get('country', '')
            
            # Billing address (mirror main address for import convenience)
            customer['Billing First Name'] = first_name
            customer['Billing Last Name'] = last_name
            customer['Billing Company'] = parsed.get('company', '')
            customer['Billing Email'] = parsed.get('email', '')
            customer['Billing Phone'] = parsed.get('tel', '')
            customer['Billing Street'] = street
            customer['Billing Street Number'] = street_num
            customer['Billing City'] = parsed.get('city', '')
            customer['Billing Zip Code'] = parsed.get('zipCode', '')
            customer['Billing Country'] = parsed.get('country', '')
            
            customers.append(customer)
    
    # Save cache
    save_cache(cache)
    
    print(f"✅ Processed {len(customers)} customers")
    print(f"   📡 API calls: {api_calls}")
    print(f"   💾 Cache hits: {cache_hits}")
    print(f"   ⏭️ Skipped: {skipped}")
    
    # Output to Excel
    if output_file.endswith('.xlsx') and HAS_OPENPYXL:
        print(f"💾 Writing Excel output to: {output_file}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        headers = [
            'Name', 'Facebook Name', 'Facebook ID', 'Facebook URL', 'Email', 'Phone',
            'Address', 'City', 'Zip Code', 'Country', 'Notes',
            'Type', 'Preferred Language', 'Preferred Currency',
            'ICO', 'DIC', 'VAT ID', 'VAT Number', 'Tax ID',
            'Billing First Name', 'Billing Last Name', 'Billing Company',
            'Billing Email', 'Billing Phone', 'Billing Street',
            'Billing Street Number', 'Billing City', 'Billing Zip Code', 'Billing Country'
        ]
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, customer in enumerate(customers, 2):
            for col_idx, header in enumerate(headers, 1):
                value = customer.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_file)
    else:
        # CSV fallback
        output_csv = output_file.replace('.xlsx', '.csv')
        print(f"💾 Writing CSV output to: {output_csv}")
        
        headers = [
            'Name', 'Facebook Name', 'Facebook ID', 'Facebook URL', 'Email', 'Phone',
            'Address', 'City', 'Zip Code', 'Country', 'Notes',
            'Type', 'Preferred Language', 'Preferred Currency',
            'ICO', 'DIC', 'VAT ID', 'VAT Number', 'Tax ID',
            'Billing First Name', 'Billing Last Name', 'Billing Company',
            'Billing Email', 'Billing Phone', 'Billing Street',
            'Billing Street Number', 'Billing City', 'Billing Zip Code', 'Billing Country'
        ]
        
        with open(output_csv, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(customers)
    
    print(f"🚀 Done! Generated {len(customers)} customers ready for import")
    
    # Show sample output
    print("\n📋 Sample output (first 10 customers with addresses):")
    print("-" * 100)
    sample = [c for c in customers if c['City'] or c['Shipping Street']][:10]
    for customer in sample:
        name = customer['Name'][:25] if customer['Name'] else '(no name)'
        city = customer['City'][:15] if customer['City'] else ''
        country = customer['Country'][:10] if customer['Country'] else ''
        fb_id = customer['Facebook ID'][:20] if customer['Facebook ID'] else ''
        print(f"  {name:25} | {city:15} | {country:10} | {fb_id}")


def main():
    # Default files
    input_file = 'attached_assets/Pasted-User-ID-Address-luutuan-1990-Luu-Tuan-Kasiano-nail-OC-V_1767446501325.txt'
    output_file = 'customers_import_ready.xlsx'
    
    # API URL from environment variable or command line
    # Set ADDRESS_PARSE_API_URL environment variable or pass as argument
    api_url = os.environ.get('ADDRESS_PARSE_API_URL')
    session_cookie = os.environ.get('SESSION_COOKIE')
    
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    if len(sys.argv) >= 4:
        api_url = sys.argv[3]
    if len(sys.argv) >= 5:
        session_cookie = sys.argv[4]
    
    if not os.path.exists(input_file):
        print(f"❌ Error: Input file not found: {input_file}")
        sys.exit(1)
    
    if api_url:
        print(f"🌐 Using API for address parsing: {api_url}")
    else:
        print("📝 Using local regex parsing (set ADDRESS_PARSE_API_URL for AI parsing)")
    
    process_customers(input_file, output_file, api_url, session_cookie)


if __name__ == '__main__':
    main()
